from __future__ import annotations

import csv
import io
import json
import os
import platform
import sqlite3
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from flask import Blueprint, Response, current_app, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from werkzeug.utils import secure_filename

from .db import ORDER_FIELDS, TIMESTAMP_FIELDS, Database, utcnow
from .imports import inspect_file, import_rows, read_csv_rows, read_xlsx_rows
from .i18n import translate
from .services import PHASES, STATUS_VALUES, TIMELINE, create_backup, dashboard, enrich_order, list_orders, restore_backup, save_order, validate_order

bp = Blueprint("main", __name__)


def db() -> Database: return current_app.extensions["db"]
def unit() -> str: return request.args.get("unit") or db().scalar("SELECT value FROM settings WHERE key='display_unit'") or "workdays"
def language() -> str:
    value = db().scalar("SELECT value FROM settings WHERE key='language'") or "de"
    return value if value in ("de", "en") else "de"
def common() -> dict[str, Any]:
    lang = language()
    return {"locations": db().all("SELECT * FROM locations WHERE active=1 ORDER BY code"), "statuses": STATUS_VALUES, "unit": unit(), "language": lang, "t": lambda text: translate(text, lang)}


@bp.route("/health")
def health():
    try:
        db().scalar("SELECT 1")
        return {"status": "ok", "database": "sqlite" if db().is_sqlite else "postgresql"}
    except Exception:
        return {"status": "error"}, 503


@bp.app_template_filter("dt")
def fmt_dt(value: Any) -> str:
    if not value: return "–"
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed.strftime("%Y-%m-%d %H:%M" if language() == "en" else "%d.%m.%Y %H:%M")
    except ValueError: return str(value)


@bp.app_template_filter("num")
def fmt_num(value: Any, digits: int = 1) -> str:
    if value is None: return "–"
    formatted = f"{float(value):,.{digits}f}"
    return formatted if language() == "en" else formatted.replace(",", "X").replace(".", ",").replace("X", ".")


@bp.route("/")
def index():
    data = dashboard(db(), {k: v for k, v in request.args.items() if v}, unit())
    return render_template("dashboard.html", data=data, title="Dashboard", **common())


@bp.route("/orders")
def orders():
    items = list_orders(db(), request.args.to_dict(), unit())
    carriers = db().all("SELECT DISTINCT carrier FROM orders WHERE carrier IS NOT NULL ORDER BY carrier")
    return render_template("orders.html", orders=items, carriers=carriers, title="Bestellungen", open_view=False, **common())


@bp.route("/orders/open")
def open_orders():
    filters = request.args.to_dict(); filters["open"] = "1"
    items = list_orders(db(), filters, unit())
    carriers = db().all("SELECT DISTINCT carrier FROM orders WHERE carrier IS NOT NULL ORDER BY carrier")
    return render_template("orders.html", orders=items, carriers=carriers, title="Offene Bestellungen", open_view=True, **common())


def form_data() -> dict[str, Any]:
    data = {k: request.form.get(k, "").strip() for k in ORDER_FIELDS}
    for field in TIMESTAMP_FIELDS:
        value = data.get(field)
        if value: data[field] = datetime.fromisoformat(value).isoformat()
    return data


@bp.route("/orders/new", methods=["GET", "POST"])
def order_new():
    item: dict[str, Any] = {}
    if request.method == "POST":
        item = form_data(); errors = validate_order(item, db())
        if not errors:
            try:
                oid = save_order(db(), item); flash("Bestellung wurde angelegt.", "success")
                return redirect(url_for("main.order_detail", order_id=oid))
            except (ValueError, sqlite3.IntegrityError) as exc: errors = [str(exc)]
        for error in errors: flash(error, "error")
    return render_template("order_form.html", order=item, title="Bestellung erfassen", reasons=db().all("SELECT * FROM delay_reasons WHERE active=1 ORDER BY name"), **common())


@bp.route("/orders/<int:order_id>")
def order_detail(order_id: int):
    row = db().one("""SELECT o.*,l.code location_code,l.name location_name,l.sla_order_ready,l.sla_total,l.standard_transport_days,d.name delay_reason
      FROM orders o JOIN locations l ON l.id=o.location_id LEFT JOIN delay_reasons d ON d.id=o.delay_reason_id WHERE o.id=?""", (order_id,))
    if not row: return ("Bestellung nicht gefunden", 404)
    item = enrich_order(row, unit())
    events = db().all("SELECT * FROM events WHERE order_id=? ORDER BY event_at,id", (order_id,))
    timeline = [{"field": f, "label": label, "value": item.get(f)} for f, label in TIMELINE]
    return render_template("order_detail.html", order=item, events=events, timeline=timeline, phases=PHASES, title=item["tracking_id"], **common())


@bp.route("/orders/<int:order_id>/edit", methods=["GET", "POST"])
def order_edit(order_id: int):
    existing = db().one("SELECT * FROM orders WHERE id=?", (order_id,))
    if not existing: return ("Bestellung nicht gefunden", 404)
    item = dict(existing)
    if request.method == "POST":
        item = form_data(); errors = validate_order(item, db(), order_id)
        if not errors:
            try:
                save_order(db(), item, order_id); flash("Änderungen wurden protokolliert.", "success")
                return redirect(url_for("main.order_detail", order_id=order_id))
            except (ValueError, sqlite3.IntegrityError) as exc: errors = [str(exc)]
        for error in errors: flash(error, "error")
    return render_template("order_form.html", order=item, title="Bestellung bearbeiten", reasons=db().all("SELECT * FROM delay_reasons WHERE active=1 ORDER BY name"), **common())


@bp.route("/comparison")
def comparison():
    data = dashboard(db(), request.args.to_dict(), unit())
    return render_template("comparison.html", data=data, title="Standortvergleich", **common())


@bp.route("/quality")
def quality():
    items = list_orders(db(), {"quality": "1", **request.args.to_dict()}, unit())
    issue_counts: dict[str, int] = {}
    for item in items:
        for issue in item["quality_issues"]: issue_counts[issue["message"]] = issue_counts.get(issue["message"], 0) + 1
    return render_template("quality.html", orders=items, issue_counts=issue_counts, title="Datenqualität", **common())


@bp.route("/import", methods=["GET", "POST"])
def import_page():
    if request.method == "POST":
        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename: flash("Bitte eine Datei auswählen.", "error"); return redirect(request.url)
        ext = Path(uploaded.filename).suffix.lower()
        if ext not in (".csv", ".tsv", ".xlsx"): flash("Erlaubt sind CSV, TSV und XLSX.", "error"); return redirect(request.url)
        token = uuid.uuid4().hex; filename = secure_filename(uploaded.filename) or f"import{ext}"
        path = Path(current_app.config["UPLOAD_DIR"]) / f"{token}{ext}"; uploaded.save(path)
        try:
            info = inspect_file(path, request.form.get("sheet") or None, int(request.form.get("header_row", 1)))
            if request.form.get("profile_id"):
                profile = db().one("SELECT mapping_json FROM import_profiles WHERE id=?", (int(request.form["profile_id"]),))
                if profile:
                    info["mapping"] = json.loads(profile["mapping_json"])
        except Exception as exc: path.unlink(missing_ok=True); flash(f"Datei konnte nicht gelesen werden: {exc}", "error"); return redirect(request.url)
        session["import"] = {"path": str(path), "filename": filename, "sheet": request.form.get("sheet") or (info["sheets"][0] if info["sheets"] else None), "header_row": int(request.form.get("header_row", 1))}
        return render_template("import_map.html", info=info, targets=[*ORDER_FIELDS, "location_code", "delay_reason"], title="Import zuordnen", **common())
    logs = db().all("SELECT * FROM import_logs ORDER BY imported_at DESC LIMIT 20")
    profiles = db().all("SELECT * FROM import_profiles ORDER BY name")
    return render_template("import.html", logs=logs, profiles=profiles, title="Import", **common())


@bp.route("/import/confirm", methods=["POST"])
def import_confirm():
    state = session.get("import")
    if not state or not Path(state["path"]).exists(): flash("Importsitzung ist abgelaufen.", "error"); return redirect(url_for("main.import_page"))
    path = Path(state["path"]); mapping = {k[4:]: v for k, v in request.form.items() if k.startswith("map_") and v}
    try:
        rows = read_xlsx_rows(path, state.get("sheet"), state.get("header_row", 1)) if path.suffix == ".xlsx" else read_csv_rows(path)
        result = import_rows(db(), rows, mapping, state["filename"], request.form.get("duplicate_mode", "skip"), Path(current_app.config["BACKUP_DIR"]))
        profile_name = request.form.get("profile_name", "").strip()
        if profile_name:
            now = utcnow()
            with db().transaction() as con: con.execute("""INSERT INTO import_profiles(name,file_type,sheet_name,header_row,mapping_json,created_at,updated_at)
              VALUES (?,?,?,?,?,?,?) ON CONFLICT(name) DO UPDATE SET mapping_json=excluded.mapping_json,sheet_name=excluded.sheet_name,header_row=excluded.header_row,updated_at=excluded.updated_at""", (profile_name, path.suffix, state.get("sheet"), state.get("header_row", 1), json.dumps(mapping, ensure_ascii=False), now, now))
        flash(f"Import abgeschlossen: {result['new']} neu, {result['updated']} aktualisiert, {result['skipped']} übersprungen.", "success" if not result["errors"] else "warning")
        for err in result["errors"][:8]: flash(err, "error")
    except Exception as exc: flash(f"Import fehlgeschlagen: {exc}", "error")
    finally: path.unlink(missing_ok=True); session.pop("import", None)
    return redirect(url_for("main.import_page"))


@bp.route("/reports")
def reports():
    data = dashboard(db(), request.args.to_dict(), unit())
    reasons = db().all("SELECT d.name,COUNT(*) count FROM orders o JOIN delay_reasons d ON d.id=o.delay_reason_id GROUP BY d.id ORDER BY count DESC LIMIT 8")
    bottleneck = max((v for k, v in data["phase_stats"].items() if k in ("order_release","release_pick","waiting_pick","pick_cycle","pack","ready_pickup","transport")), key=lambda x: x["median"] or -1, default={"label":"–"})
    return render_template("reports.html", data=data, reasons=reasons, bottleneck=bottleneck, title="Management-Bericht", **common())


def export_rows() -> list[dict[str, Any]]: return list_orders(db(), request.args.to_dict(), unit())
EXPORT_COLUMNS = [("tracking_id","Tracking-ID"),("location_code","Standort"),("order_at","Bestelldatum"),("at_sales_order_no","AT-Auftrag"),("status","Status"),("carrier","Spedition"),("pallet_count","Paletten"),("delay_reason","Verzögerungsgrund")]


@bp.route("/export/orders.csv")
def export_csv():
    output = io.StringIO(); writer = csv.writer(output, delimiter=";"); writer.writerow([label for _, label in EXPORT_COLUMNS] + ["Order-to-Ready", "Transport", "Gesamt", "SLA-Status"])
    for o in export_rows(): writer.writerow([o.get(k) for k, _ in EXPORT_COLUMNS] + [o["durations"]["order_ready"],o["durations"]["transport"],o["durations"]["total"],o["sla_state"]])
    return Response("\ufeff" + output.getvalue(), mimetype="text/csv", headers={"Content-Disposition":"attachment; filename=bestellungen.csv"})


@bp.route("/export/orders.xlsx")
def export_xlsx():
    wb = Workbook(); ws = wb.active; ws.title = "Bestellungen"; ws.append([label for _, label in EXPORT_COLUMNS] + ["Order-to-Ready", "Transport", "Gesamt", "SLA-Status"])
    for o in export_rows(): ws.append([o.get(k) for k, _ in EXPORT_COLUMNS] + [o["durations"]["order_ready"],o["durations"]["transport"],o["durations"]["total"],o["sla_state"]])
    for cell in ws[1]: cell.font = __import__("openpyxl").styles.Font(bold=True); cell.fill = __import__("openpyxl").styles.PatternFill("solid", fgColor="DCE8F2")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return send_file(stream, as_attachment=True, download_name="bestellungen.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/master", methods=["GET", "POST"])
def master():
    if request.method == "POST":
        action = request.form.get("action")
        try:
            with db().transaction() as con:
                if action == "location":
                    con.execute("""INSERT INTO locations(code,name,country,standard_transport_days,sla_order_ready,sla_total,notes)
                      VALUES (?,?,?,?,?,?,?)""", (request.form["code"].strip().upper(),request.form["name"].strip(),request.form["country"].strip(),float(request.form.get("standard_transport_days") or 0),float(request.form.get("sla_order_ready") or 0),float(request.form.get("sla_total") or 0),request.form.get("notes","")))
                elif action == "reason": con.execute("INSERT INTO delay_reasons(name,category) VALUES (?,?)", (request.form["name"].strip(), request.form.get("category","intern")))
            flash("Stammdaten gespeichert.", "success")
        except (ValueError, sqlite3.IntegrityError) as exc: flash(f"Speichern nicht möglich: {exc}", "error")
        return redirect(request.url)
    return render_template("master.html", all_locations=db().all("SELECT * FROM locations ORDER BY code"), reasons=db().all("SELECT * FROM delay_reasons ORDER BY name"), title="Stammdaten", **common())


@bp.route("/settings", methods=["GET", "POST"])
def settings():
    if request.method == "POST":
        now = utcnow()
        with db().transaction() as con:
            for key in ("warning_no_pick_days","warning_pick_idle_days","warning_ready_pickup_days","display_unit","language"):
                value = request.form.get(key, "")
                if key == "language" and value not in ("de", "en"): value = "de"
                con.execute("INSERT INTO settings(key,value,updated_at) VALUES (?,?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=excluded.updated_at", (key, value, now))
            cfg = {key: request.form.get(key,"").strip() for key in ("system_type","base_url","tenant","environment","company","client_id","scope","api_version")}
            con.execute("UPDATE settings SET value=?,updated_at=? WHERE key='connector_config'", (json.dumps(cfg), now))
        lang = request.form.get("language") if request.form.get("language") in ("de", "en") else "de"
        flash(translate("Einstellungen gespeichert. Zugangsdaten werden bewusst nicht gespeichert.", lang), "success"); return redirect(request.url)
    values = {r["key"]:r["value"] for r in db().all("SELECT * FROM settings")}; connector = json.loads(values.get("connector_config","{}"))
    return render_template("settings.html", values=values, connector=connector, title="Einstellungen", **common())


@bp.route("/settings/test-connection", methods=["POST"])
def test_connection():
    cfg = json.loads(db().scalar("SELECT value FROM settings WHERE key='connector_config'") or "{}")
    missing = [key for key in (["base_url", "company"] if cfg.get("system_type") != "Datei" else []) if not cfg.get(key)]
    if missing:
        flash("Konfiguration unvollständig: " + ", ".join(missing), "error")
    else:
        flash("Konfiguration ist syntaktisch vollständig. Der Live-Test wird nach Freigabe der konkreten API-/OData-Endpunkte aktiviert.", "warning")
    return redirect(url_for("main.settings"))


@bp.route("/system", methods=["GET", "POST"])
def system_info():
    if request.method == "POST" and request.form.get("action") == "backup":
        path = create_backup(db(), Path(current_app.config["BACKUP_DIR"])); flash(f"Backup {path.name} erstellt.", "success"); return redirect(request.url)
    backups = db().all("SELECT * FROM backups ORDER BY created_at DESC")
    info = {"python": sys.version.split()[0], "platform": platform.platform(), "database": str(db().path), "database_size": db().path.stat().st_size if db().path.exists() else 0, "orders": db().scalar("SELECT COUNT(*) FROM orders"), "sqlite": sqlite3.sqlite_version}
    return render_template("system.html", info=info, backups=backups, title="Systeminformationen", **common())


@bp.route("/system/backup/<path:filename>")
def download_backup(filename: str):
    safe = secure_filename(filename); path = Path(current_app.config["BACKUP_DIR"]) / safe
    if not path.exists(): return ("Backup nicht gefunden", 404)
    return send_file(path, as_attachment=True, download_name=safe)


@bp.route("/system/restore", methods=["POST"])
def restore():
    uploaded = request.files.get("backup")
    if not uploaded or Path(uploaded.filename or "").suffix.lower() not in (".sqlite3", ".db"): flash("Bitte eine SQLite-Sicherung auswählen. Cloud-Restore erfolgt direkt über Supabase.", "error"); return redirect(url_for("main.system_info"))
    temp = Path(current_app.config["UPLOAD_DIR"]) / f"restore-{uuid.uuid4().hex}.sqlite3"; uploaded.save(temp)
    try: restore_backup(db(), temp, Path(current_app.config["BACKUP_DIR"])); flash("Sicherung wurde wiederhergestellt. Die vorherige Datenbank wurde automatisch gesichert.", "success")
    except Exception as exc: flash(f"Wiederherstellung abgebrochen: {exc}", "error")
    finally: temp.unlink(missing_ok=True)
    return redirect(url_for("main.system_info"))
