from __future__ import annotations

import csv
import io
import json
import re
import sqlite3
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .db import Database, ORDER_FIELDS, utcnow
from .services import create_backup, parse_dt, save_order, validate_order


ALIASES = {
 "tracking_id": ["tracking id", "tracking-id", "trackingid"],
 "warehouse_order_no": ["aussenlager-bestellnummer", "außenlager-bestellnummer", "purchase order no", "po number", "bestellnr", "bestellnummer"],
 "at_sales_order_no": ["at-verkaufsauftragsnummer", "verkaufsauftragsnummer", "sales order no", "sales order number", "auftragsnr", "auftragsnr."],
 "external_document_no": ["externe belegnummer", "external document no", "external document number", "ch-bestellnummer"],
 "location_code": ["standort", "standortcode", "location", "location code", "land", "country"],
 "order_at": ["bestelldatum", "order date", "order datetime"],
 "created_at_at": ["erstellungsdatum", "created at", "systemcreatedat", "auftrag erstellt"],
 "released_at": ["freigabedatum", "release date", "released at"],
 "pick_created_at": ["pick-erstellungsdatum", "pick created", "warehouse pick created"],
 "pick_started_at": ["pick-startdatum", "pick started", "assignment date"],
 "pick_registered_at": ["pick-registrierungsdatum", "registered pick date", "pick registered"],
 "packed_at": ["verpackung abgeschlossen", "packed at", "packing complete"],
 "documents_complete_at": ["dokumente vollständig", "documents complete", "invoice ready"],
 "ready_at": ["abholbereit", "ready for pickup", "ready at"],
 "picked_up_at": ["abholdatum", "pickup date", "picked up at", "shipment date"],
 "shipment_posted_at": ["versandbuchungsdatum", "shipment posting date", "posting date"],
 "received_at": ["wareneingangsdatum", "receipt date", "received at"],
 "line_count": ["anzahl positionen", "lines", "line count"],
 "total_quantity": ["gesamtmenge", "total quantity", "quantity"],
 "pallet_count": ["anzahl paletten", "pallets", "pallet count"],
 "total_weight": ["gesamtgewicht", "total weight", "weight"],
 "order_value": ["auftragswert", "order value", "amount"],
 "carrier": ["spedition", "carrier", "forwarder"],
 "transport_mode": ["transportart", "transport mode"],
 "priority": ["priorität", "priority"], "responsible": ["verantwortliche person", "responsible"],
 "notes": ["bemerkung", "notes", "comment"], "delay_reason": ["verzögerungsgrund", "delay reason"],
}


def norm(text: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^\w\s-]", "", str(text or "").strip().lower().replace("_", " ")))


def suggest_mapping(headers: Iterable[Any]) -> dict[str, str]:
    targets = {norm(alias): field for field, aliases in ALIASES.items() for alias in [field, *aliases]}
    return {str(header): targets[norm(header)] for header in headers if norm(header) in targets}


def sniff_delimiter(text: str) -> str:
    try: return csv.Sniffer().sniff(text[:4096], delimiters=";,\t|").delimiter
    except csv.Error: return ";"


def decode_csv(raw: bytes) -> str:
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try: return raw.decode(enc)
        except UnicodeDecodeError: continue
    raise ValueError("CSV-Zeichencodierung konnte nicht erkannt werden.")


def read_csv_rows(path: Path, delimiter: str | None = None) -> list[dict[str, Any]]:
    text = decode_csv(path.read_bytes())
    return list(csv.DictReader(io.StringIO(text), delimiter=delimiter or sniff_delimiter(text)))


def workbook_sheets(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try: return wb.sheetnames
    finally: wb.close()


def read_xlsx_rows(path: Path, sheet_name: str | None = None, header_row: int = 1) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        for _ in range(max(0, header_row - 1)): next(rows, None)
        headers = [str(v).strip() if v is not None else f"Spalte {i+1}" for i, v in enumerate(next(rows, []))]
        return [dict(zip(headers, vals)) for vals in rows if any(v is not None for v in vals)]
    finally: wb.close()


def inspect_file(path: Path, sheet_name: str | None = None, header_row: int = 1) -> dict[str, Any]:
    if path.suffix.lower() == ".xlsx":
        sheets = workbook_sheets(path); rows = read_xlsx_rows(path, sheet_name or sheets[0], header_row)
    elif path.suffix.lower() in (".csv", ".tsv"):
        sheets = []; rows = read_csv_rows(path)
    else: raise ValueError("Nur CSV, TSV und XLSX werden unterstützt.")
    headers = list(rows[0]) if rows else []
    return {"sheets": sheets, "headers": headers, "mapping": suggest_mapping(headers), "preview": rows[:10], "row_count": len(rows)}


def decimal_value(value: Any) -> float | None:
    if value in (None, ""): return None
    if isinstance(value, (int, float)): return float(value)
    text = str(value).strip().replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
    elif "," in text: text = text.replace(",", ".")
    return float(text)


def normalize_row(raw: dict[str, Any], mapping: dict[str, str], db: Database) -> dict[str, Any]:
    out: dict[str, Any] = {target: raw.get(source) for source, target in mapping.items() if target}
    for key, value in list(out.items()):
        if isinstance(value, (datetime, date)): out[key] = datetime.combine(value, datetime.min.time(), tzinfo=timezone.utc).isoformat() if isinstance(value, date) and not isinstance(value, datetime) else value.replace(tzinfo=value.tzinfo or timezone.utc).isoformat()
        elif key.endswith("_at") and value not in (None, ""):
            parsed = parse_dt(value)
            if not parsed: raise ValueError(f"Ungültiges Datum für {key}: {value}")
            out[key] = parsed.isoformat()
    for key in ("total_quantity", "total_weight", "order_value"):
        if key in out: out[key] = decimal_value(out[key])
    for key in ("line_count", "pallet_count"):
        if out.get(key) not in (None, ""): out[key] = int(decimal_value(out[key]) or 0)
    location_code = str(out.pop("location_code", "")).strip()
    location = db.one("SELECT id FROM locations WHERE code=? COLLATE NOCASE OR name=? COLLATE NOCASE", (location_code, location_code))
    if not location: raise ValueError(f"Unbekannter Standort: {location_code}")
    out["location_id"] = location["id"]
    reason = str(out.pop("delay_reason", "") or "").strip()
    if reason:
        row = db.one("SELECT id FROM delay_reasons WHERE name=? COLLATE NOCASE", (reason,))
        if not row:
            with db.transaction() as con: rid = con.execute("INSERT INTO delay_reasons(name,category) VALUES (?,?)", (reason, "importiert")).lastrowid
            out["delay_reason_id"] = rid
        else: out["delay_reason_id"] = row["id"]
    if not out.get("tracking_id"):
        out["tracking_id"] = out.get("at_sales_order_no") or out.get("warehouse_order_no") or out.get("external_document_no")
    out.setdefault("priority", "normal")
    out["data_source"] = "Dateiimport"
    return out


def find_duplicate(db: Database, row: dict[str, Any]) -> int | None:
    clauses, params = [], []
    for field in ("tracking_id", "warehouse_order_no", "at_sales_order_no", "external_document_no"):
        if row.get(field): clauses.append(f"{field}=? COLLATE NOCASE"); params.append(str(row[field]))
    if not clauses: return None
    found = db.one("SELECT id FROM orders WHERE " + " OR ".join(clauses) + " LIMIT 1", tuple(params))
    return found["id"] if found else None


def import_rows(db: Database, rows: list[dict[str, Any]], mapping: dict[str, str], filename: str, duplicate_mode: str = "skip", backup_dir: Path | None = None) -> dict[str, Any]:
    if backup_dir and rows: create_backup(db, backup_dir, f"Vor Import {filename}")
    result = {"read": len(rows), "new": 0, "updated": 0, "skipped": 0, "errors": [], "warnings": []}
    for number, raw in enumerate(rows, 2):
        try:
            item = normalize_row(raw, mapping, db); existing = find_duplicate(db, item)
            if existing and duplicate_mode == "skip": result["skipped"] += 1; continue
            if existing:
                current = dict(db.one("SELECT * FROM orders WHERE id=?", (existing,)))
                if duplicate_mode == "fill": item = {**current, **{k: v for k, v in item.items() if current.get(k) in (None, "")}}
                elif duplicate_mode != "update": result["skipped"] += 1; result["warnings"].append(f"Zeile {number}: Duplikat zur manuellen Prüfung"); continue
                errors = validate_order(item, db, existing)
                if errors: raise ValueError("; ".join(errors))
                save_order(db, item, existing, "Dateiimport"); result["updated"] += 1
            else:
                errors = validate_order(item, db)
                if errors: raise ValueError("; ".join(errors))
                save_order(db, item, source="Dateiimport"); result["new"] += 1
        except (ValueError, sqlite3.IntegrityError) as exc:
            result["skipped"] += 1; result["errors"].append(f"Zeile {number}: {exc}")
    with db.transaction() as con:
        con.execute("""INSERT INTO import_logs(filename,file_type,imported_at,rows_read,rows_new,rows_updated,rows_skipped,errors,warnings)
          VALUES (?,?,?,?,?,?,?,?,?)""", (filename, Path(filename).suffix.lower(), utcnow(), result["read"], result["new"], result["updated"], result["skipped"], json.dumps(result["errors"], ensure_ascii=False), json.dumps(result["warnings"], ensure_ascii=False)))
    return result
