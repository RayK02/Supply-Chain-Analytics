from __future__ import annotations

import base64
import io
import re
import zipfile
from datetime import date, datetime
from typing import Any, BinaryIO, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.inventory_utils import (
    DEFAULT_PARAMETERS,
    PARAMETER_EXPORT_NAMES,
    as_float,
    display_article_key,
    normalize_article_key,
    parse_code_set,
)

SHEETS = {
    "sales": "Artikelposten",
    "articles": "Artikel_Stamm",
    "current": "Lagerhaltungsdaten_IST",
    "current_export": "Lagerhaltungsdatenübersicht",
    "vpe": "VPE",
    "params": "Parameter",
}

ALIASES = {
    "article": ["Artikelnummer", "Artikelnr.", "Artikelnr", "Artikel-Nr.", "Nr.", "Nr", "Item No.", "Item No"],
    "variant": ["Variantencode", "Variant Code"],
    "procurement": ["Beschaffungsmethode", "Replenishment System", "Procurement Method"],
    "description": ["Beschreibung", "Artikelbeschreibung", "Description"],
    "date": ["Buchungsdatum", "Posting Date", "Datum"],
    "type": ["Belegart", "Document Type"],
    "qty": ["Menge", "Quantity"],
    "location": ["Lagerortcode", "Lagerort", "Location Code"],
    "vpe": ["VPE", "Verpackungseinheit", "Menge je VPE", "Qty. per Unit of Measure"],
    "stock": ["Lagerbestand", "Bestand", "Inventory", "Current Inventory"],
    "minimum": ["Mindestbestand", "Minimalbestand", "Minimum Inventory", "Sicherheitsbestand"],
    "order": ["Bestellmenge", "Order Quantity", "Bestelllosgröße", "Bestelllosgroesse"],
    "maximum": ["Maximalbestand", "Maximum Inventory"],
}

MAX_SHEET_ROWS = 100_000
MAX_ZIP_ENTRIES = 500
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_COMPRESSION_RATIO = 200

PARAMETER_EXPORT_ORDER = [
    "document_type",
    "return_document_types",
    "months_average",
    "xyz_months",
    "analysis_start_date",
    "analysis_end_date",
    "abc_a_threshold",
    "abc_b_threshold",
    "xyz_x_threshold",
    "xyz_y_threshold",
    "xyz_min_months",
    "minimum_factor_a",
    "minimum_factor_b",
    "minimum_factor_c",
    "order_interval_a",
    "order_interval_b",
    "order_interval_c",
    "automatic_locations",
    "manual_locations",
    "max_import_rows",
]


def canon(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"[^a-z0-9]+", "", text)


def _rewind(file_object: BinaryIO) -> None:
    try:
        file_object.seek(0)
    except (AttributeError, OSError):
        pass


def validate_xlsx_archive(file_object: BinaryIO, filename: str = "Arbeitsmappe") -> None:
    _rewind(file_object)
    try:
        with zipfile.ZipFile(file_object) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ZIP_ENTRIES:
                raise ValueError(f"{filename}: Zu viele Bestandteile in der XLSX-Datei.")
            total_uncompressed = sum(entry.file_size for entry in entries)
            if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
                raise ValueError(f"{filename}: Entpackter Inhalt ist zu gross.")
            for entry in entries:
                if entry.flag_bits & 0x1:
                    raise ValueError(f"{filename}: Verschlüsselte XLSX-Dateien werden nicht unterstützt.")
                if entry.compress_size > 0 and entry.file_size > 1_000_000:
                    ratio = entry.file_size / entry.compress_size
                    if ratio > MAX_COMPRESSION_RATIO:
                        raise ValueError(f"{filename}: Ungewöhnlich hohe ZIP-Kompression erkannt.")
    except zipfile.BadZipFile as exc:
        raise ValueError(f"{filename}: Keine gültige XLSX-Datei.") from exc
    finally:
        _rewind(file_object)


def sheet(workbook, name: str, required: bool = True) -> str | None:
    found = {canon(value): value for value in workbook.sheetnames}.get(canon(name))
    if not found and required:
        raise ValueError(f"Tabellenblatt fehlt: {name}")
    return found


def effective_row_limit(parameters: Mapping[str, Any] | None) -> int:
    raw = as_float((parameters or {}).get("max_import_rows"))
    if raw is None:
        return MAX_SHEET_ROWS
    if not float(raw).is_integer():
        raise ValueError("Maximale Importzeilen müssen eine ganze Zahl sein.")
    limit = int(raw)
    if not 1 <= limit <= MAX_SHEET_ROWS:
        raise ValueError(f"Maximale Importzeilen müssen zwischen 1 und {MAX_SHEET_ROWS:,} liegen.")
    return limit


def rows(workbook, name: str, max_rows: int = MAX_SHEET_ROWS) -> tuple[list[str], list[dict[str, Any]]]:
    worksheet = workbook[name]
    iterator = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(iterator)
    except StopIteration:
        return [], []

    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        base = str(value).strip() if value is not None else f"Spalte_{index + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")

    data: list[dict[str, Any]] = []
    for row_number, row in enumerate(iterator, start=2):
        if row_number > max_rows + 1:
            raise ValueError(f"Tabellenblatt {name} überschreitet {max_rows:,} Datenzeilen.")
        if not any(value not in (None, "") for value in row):
            continue
        data.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
    return headers, data


def col(headers: list[str], key: str, required: bool = True) -> str | None:
    by_canonical = {canon(value): value for value in headers}
    for alias in ALIASES[key]:
        found = by_canonical.get(canon(alias))
        if found:
            return found
    if required:
        raise ValueError(f"Spalte fehlt: {ALIASES[key][0]}")
    return None


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for date_format in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), date_format).date()
            except ValueError:
                pass
    return None


def _parameter_number(value: Any) -> float | None:
    if isinstance(value, str) and value.strip().endswith("%"):
        parsed = as_float(value.strip()[:-1])
        return parsed / 100 if parsed is not None else None
    return as_float(value)


def _parameter_mapping() -> dict[str, str]:
    aliases: dict[str, list[str]] = {
        "document_type": ["Belegart", "Verkaufsbelegart", "document_type"],
        "return_document_types": [
            "Rückgabebelegarten", "Rücklaufbelegarten", "Gutschriftbelegarten", "return_document_types"
        ],
        "months_average": ["Monate", "Durchschnittsmonate", "months_average"],
        "xyz_months": ["XYZ Monate", "XYZ-Zeitraum", "xyz_months"],
        "analysis_start_date": ["Startdatum", "Startdatum Analyse", "analysis_start_date"],
        "analysis_end_date": ["Enddatum", "Enddatum Analyse", "analysis_end_date"],
        "abc_a_threshold": ["ABC A Grenze", "abc_a_threshold"],
        "abc_b_threshold": ["ABC B Grenze", "abc_b_threshold"],
        "xyz_x_threshold": ["XYZ X Grenze", "xyz_x_threshold"],
        "xyz_y_threshold": ["XYZ Y Grenze", "xyz_y_threshold"],
        "xyz_min_months": ["XYZ Mindestmonate", "xyz_min_months"],
        "minimum_factor_a": ["Mindestfaktor A", "minimum_factor_a"],
        "minimum_factor_b": ["Mindestfaktor B", "minimum_factor_b"],
        "minimum_factor_c": ["Mindestfaktor C", "minimum_factor_c"],
        "order_interval_a": ["Bestellintervall A", "order_interval_a"],
        "order_interval_b": ["Bestellintervall B", "order_interval_b"],
        "order_interval_c": ["Bestellintervall C", "order_interval_c"],
        "automatic_locations": ["Automatische Lagerorte", "automatic_locations"],
        "manual_locations": ["Manuelle Lagerorte", "manual_locations"],
        "max_import_rows": ["Maximale Importzeilen", "Max Import Rows", "max_import_rows"],
    }
    return {canon(alias): key for key, values in aliases.items() for alias in values}


def params(workbook) -> dict[str, Any]:
    output: dict[str, Any] = dict(DEFAULT_PARAMETERS)
    name = sheet(workbook, SHEETS["params"], False)
    if not name:
        output["_parameter_source"] = "Standardwerte"
        output["_parameter_recognized"] = ""
        output["_parameter_ignored"] = ""
        return output

    mapping = _parameter_mapping()
    text_keys = {"document_type", "return_document_types", "automatic_locations", "manual_locations"}
    date_keys = {"analysis_start_date", "analysis_end_date"}
    recognized: list[str] = []
    ignored: list[str] = []

    for row in workbook[name].iter_rows(values_only=True):
        if not row or row[0] in (None, ""):
            continue
        raw_key = str(row[0]).strip()
        canonical_key = canon(raw_key)
        key = mapping.get(canonical_key)
        value = row[1] if len(row) > 1 else None
        if not key:
            if canonical_key not in {"parameter", "name", "bezeichnung"}:
                ignored.append(raw_key)
            continue
        if key in text_keys:
            if value not in (None, ""):
                output[key] = str(value).strip()
                recognized.append(raw_key)
            else:
                ignored.append(raw_key)
        elif key in date_keys:
            parsed_date = as_date(value)
            if parsed_date is not None:
                output[key] = parsed_date
                recognized.append(raw_key)
            elif value in (None, ""):
                output.pop(key, None)
                recognized.append(raw_key)
            else:
                ignored.append(raw_key)
        else:
            parsed = _parameter_number(value)
            if parsed is not None:
                output[key] = parsed
                recognized.append(raw_key)
            else:
                ignored.append(raw_key)

    effective_row_limit(output)
    output["_parameter_source"] = f"Excel: {name}"
    output["_parameter_recognized"] = ", ".join(recognized)
    output["_parameter_ignored"] = ", ".join(ignored)
    return output


def article_names(workbook, parameters: Mapping[str, Any] | None = None) -> dict[str, str]:
    name = sheet(workbook, SHEETS["articles"], False)
    if not name:
        return {}
    headers, data = rows(workbook, name, effective_row_limit(parameters))
    article_column = col(headers, "article")
    description_column = col(headers, "description", False)
    output: dict[str, str] = {}
    for row in data:
        key = normalize_article_key(row.get(article_column))
        if key:
            output[key] = str(row.get(description_column) or "").strip() if description_column else ""
    return output


def vpe_values(workbook, parameters: Mapping[str, Any] | None = None) -> dict[str, float]:
    name = sheet(workbook, SHEETS["vpe"], False)
    if not name:
        return {}
    headers, data = rows(workbook, name, effective_row_limit(parameters))
    article_column = col(headers, "article")
    vpe_column = col(headers, "vpe")
    output: dict[str, float] = {}
    for row in data:
        key = normalize_article_key(row.get(article_column))
        value = as_float(row.get(vpe_column))
        if key and value and value > 0:
            output[key] = value
    return output


def _current_sheet_candidates(workbook, *, scan_all_sheets: bool) -> list[str]:
    ordered: list[str] = []
    for preferred in (SHEETS["current_export"], SHEETS["current"]):
        found = sheet(workbook, preferred, False)
        if found and found not in ordered:
            ordered.append(found)
    if scan_all_sheets:
        excluded = {canon(SHEETS[key]) for key in ("sales", "articles", "vpe", "params")}
        for name in workbook.sheetnames:
            if name not in ordered and canon(name) not in excluded:
                ordered.append(name)
    return ordered


def _location_priority(location: str, automatic: set[str], manual: set[str]) -> int:
    if location in automatic:
        return 0
    if location in manual:
        return 1
    return 2


def current_values(
    workbook,
    parameters: dict[str, Any],
    *,
    source_file: str,
    required: bool = False,
    scan_all_sheets: bool = False,
) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    automatic = parse_code_set(parameters.get("automatic_locations"))
    manual = parse_code_set(parameters.get("manual_locations"))
    row_limit = effective_row_limit(parameters)
    output: dict[str, dict[str, Any]] = {}
    recognized_sheets: list[str] = []
    parsed_rows = 0
    invalid_article_rows = 0

    for sheet_name in _current_sheet_candidates(workbook, scan_all_sheets=scan_all_sheets):
        headers, data = rows(workbook, sheet_name, row_limit)
        if not headers:
            continue
        article_column = col(headers, "article", False)
        if not article_column:
            continue
        location_column = col(headers, "location", False)
        stock_column = col(headers, "stock", False)
        minimum_column = col(headers, "minimum", False)
        order_column = col(headers, "order", False)
        maximum_column = col(headers, "maximum", False)
        description_column = col(headers, "description", False)
        variant_column = col(headers, "variant", False)
        procurement_column = col(headers, "procurement", False)
        if not any((stock_column, minimum_column, order_column, maximum_column)):
            continue

        recognized_sheets.append(sheet_name)
        parsed_rows += len(data)
        for row in data:
            raw_article = row.get(article_column)
            article_key = normalize_article_key(raw_article)
            article_display = display_article_key(raw_article)
            if not article_key or not article_display:
                invalid_article_rows += 1
                continue
            location = str(row.get(location_column) or "").strip().upper() if location_column else ""
            priority = _location_priority(location, automatic, manual)
            value = {
                "article_display_current": article_display,
                "location": location,
                "manual_review": location in manual,
                "stock_current": as_float(row.get(stock_column)) if stock_column else None,
                "minimum_stock_current": as_float(row.get(minimum_column)) if minimum_column else None,
                "order_quantity_current": as_float(row.get(order_column)) if order_column else None,
                "maximum_stock_current": as_float(row.get(maximum_column)) if maximum_column else None,
                "description_current": str(row.get(description_column) or "").strip() if description_column else "",
                "variant_code_current": str(row.get(variant_column) or "").strip() if variant_column else "",
                "procurement_method_current": str(row.get(procurement_column) or "").strip() if procurement_column else "",
                "current_source_file": source_file,
                "current_source_sheet": sheet_name,
                "_priority": priority,
            }
            existing = output.get(article_key)
            if existing is None or priority <= existing["_priority"]:
                output[article_key] = value

    if required and not recognized_sheets:
        expected = "Artikelnr., Lagerortcode, Lagerbestand, Minimalbestand, Bestellmenge und Maximalbestand"
        raise ValueError(f"{source_file}: Keine erkennbare Lagerhaltungsdatenliste gefunden. Erwartete Spalten: {expected}.")

    return output, {
        "filename": source_file,
        "sheets": recognized_sheets,
        "rows": parsed_rows,
        "articles": len(output),
        "invalid_article_rows": invalid_article_rows,
    }


def merge_current_values(
    target: dict[str, dict[str, Any]],
    incoming: dict[str, dict[str, Any]],
    *,
    override_equal_priority: bool,
) -> None:
    for article_key, value in incoming.items():
        existing = target.get(article_key)
        if existing is None:
            target[article_key] = value
            continue
        incoming_priority = value.get("_priority", 99)
        existing_priority = existing.get("_priority", 99)
        if incoming_priority < existing_priority or (override_equal_priority and incoming_priority == existing_priority):
            target[article_key] = value


def sales_values(
    workbook,
    parameters: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, str], dict[str, Any]]:
    headers, data = rows(workbook, sheet(workbook, SHEETS["sales"]), effective_row_limit(parameters))
    article_column = col(headers, "article")
    description_column = col(headers, "description", False)
    date_column = col(headers, "date")
    type_column = col(headers, "type")
    quantity_column = col(headers, "qty")
    output: list[dict[str, Any]] = []
    names: dict[str, str] = {}
    invalid_article_rows = 0
    invalid_date_rows = 0
    invalid_quantity_rows = 0

    for row in data:
        raw_article = row.get(article_column)
        article_key = normalize_article_key(raw_article)
        article_display = display_article_key(raw_article)
        day = as_date(row.get(date_column))
        quantity = as_float(row.get(quantity_column))
        description = str(row.get(description_column) or "").strip() if description_column else ""
        if not article_key or not article_display:
            invalid_article_rows += 1
            continue
        if description:
            names.setdefault(article_key, description)
        if not day:
            invalid_date_rows += 1
            continue
        if quantity is None:
            invalid_quantity_rows += 1
            continue
        output.append({
            "article_key": article_key,
            "article_display": article_display,
            "description": description,
            "booking_date": day,
            "document_type": str(row.get(type_column) or "").strip(),
            "quantity": quantity,
        })

    return output, names, {
        "sheet": SHEETS["sales"],
        "rows": len(data),
        "accepted_rows": len(output),
        "invalid_article_rows": invalid_article_rows,
        "invalid_date_rows": invalid_date_rows,
        "invalid_quantity_rows": invalid_quantity_rows,
        "effective_row_limit": effective_row_limit(parameters),
    }


def compare(current: float | None, proposed: float | None) -> tuple[str, float | None]:
    if current is None or proposed is None:
        return "unknown", None
    delta = float(proposed) - float(current)
    return ("ok" if abs(delta) < 1e-9 else "increase" if delta > 0 else "decrease"), delta


def _load_current_uploads(
    uploads: Iterable[tuple[str, BinaryIO]],
    parameters: dict[str, Any],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    merged: dict[str, dict[str, Any]] = {}
    infos: list[dict[str, Any]] = []
    for filename, file_object in uploads:
        validate_xlsx_archive(file_object, filename)
        workbook = load_workbook(file_object, data_only=True, read_only=True)
        values, info = current_values(
            workbook,
            parameters,
            source_file=filename,
            required=True,
            scan_all_sheets=True,
        )
        merge_current_values(merged, values, override_equal_priority=True)
        infos.append(info)
    return merged, infos


def analyse_workbook(
    file_object: BinaryIO,
    current_file_objects: Iterable[tuple[str, BinaryIO]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    from analysis_service import analyse_workbook as service_analyse_workbook

    return service_analyse_workbook(file_object, current_file_objects)


EXPORT = [
    ("article_display", "Artikelnummer"),
    ("description", "Beschreibung"),
    ("abc_class", "ABC"),
    ("xyz_class", "XYZ"),
    ("abcxyz", "ABCXYZ"),
    ("sales_gross", "Bruttoabgang"),
    ("returns_total", "Rückläufe netto"),
    ("sales_total", "Nettoabsatz"),
    ("average_month", "Ø Absatz/Monat"),
    ("xyz_average_month", "Ø Absatz/Monat XYZ"),
    ("stock_current", "Lagerbestand IST"),
    ("minimum_stock", "Mindestbestand Vorschlag"),
    ("minimum_stock_current", "Minimalbestand IST"),
    ("minimum_stock_delta", "Delta Mindestbestand"),
    ("weekly_need", "Wochenbedarf"),
    ("order_interval", "Bestellintervall Wochen"),
    ("vpe", "VPE"),
    ("order_quantity", "Bestellmenge Vorschlag"),
    ("order_quantity_current", "Bestellmenge IST"),
    ("order_quantity_delta", "Delta Bestellmenge"),
    ("maximum_stock_current", "Maximalbestand IST"),
    ("location", "Lagerort"),
    ("variant_code_current", "Variantencode IST"),
    ("procurement_method_current", "Beschaffungsmethode IST"),
    ("current_source_file", "IST-Quelldatei"),
    ("validation_text", "Plausibilitätsprüfung"),
    ("overall_status", "Status"),
]


def _excel_safe(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    stripped = value.lstrip(" \t\r\n\ufeff")
    if stripped.startswith(("=", "+", "-", "@")) or value.startswith(("\t", "\r", "\n")):
        return "'" + value
    return value


def build_export(results: list[dict[str, Any]], meta: dict[str, Any]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ABC_Analyse"
    worksheet.append([label for _, label in EXPORT])
    for result in results:
        worksheet.append([_excel_safe(result.get(key)) for key, _ in EXPORT])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    summary = workbook.create_sheet("Zusammenfassung")
    summary.append(["Kennzahl", "Wert"])
    summary.append(["Analyse von", meta.get("analysis_start")])
    summary.append(["Analyse bis", meta.get("analysis_end")])
    summary.append(["Durchschnitt von", meta.get("average_start")])
    summary.append(["Durchschnitt bis", meta.get("average_end")])
    summary.append(["Durchschnittsmonate angefordert", meta.get("months_average_requested")])
    summary.append(["Durchschnittsmonate verwendet", meta.get("months_average_used")])
    summary.append(["XYZ von", meta.get("xyz_start")])
    summary.append(["XYZ bis", meta.get("xyz_end")])
    summary.append(["XYZ-Monate angefordert", meta.get("xyz_months_requested")])
    summary.append(["XYZ-Monate verwendet", meta.get("xyz_months_used")])
    summary.append(["Bruttoabgang", meta.get("overall_gross_sales")])
    summary.append(["Rückläufe netto", meta.get("overall_returns")])
    summary.append(["Nettoabsatz", meta.get("overall_sales")])
    summary.append(["IST-Artikel eingelesen", meta.get("current_articles")])
    summary.append(["IST-Artikel zugeordnet", meta.get("matched_current_articles")])
    summary.append(["Zusätzliche IST-Dateien", meta.get("external_current_files")])
    summary.append(["VPE fehlt", meta.get("missing_vpe_count")])
    for key, value in sorted((meta.get("counts") or {}).items()):
        summary.append([f"Anzahl {key}", value])

    sources = workbook.create_sheet("IST_Quellen")
    sources.append(["Datei", "Blätter", "Datenzeilen", "Artikel", "Ungültige Artikelzeilen"])
    for source in meta.get("current_sources") or []:
        sources.append([
            _excel_safe(source.get("filename")),
            _excel_safe(", ".join(source.get("sheets") or [])),
            source.get("rows"),
            source.get("articles"),
            source.get("invalid_article_rows"),
        ])

    import_sheet = workbook.create_sheet("Importbericht")
    import_sheet.append(["Prüfung", "Wert"])
    for key, value in (meta.get("sales_import") or {}).items():
        import_sheet.append([_excel_safe(key), _excel_safe(value)])
    for warning in meta.get("warnings") or []:
        import_sheet.append(["Warnung", _excel_safe(warning)])

    parameter_sheet = workbook.create_sheet("Parameter")
    parameter_sheet.append(["Parameter", "Wert", "Technischer Schlüssel"])
    parameters = meta.get("parameters") or {}
    for key in PARAMETER_EXPORT_ORDER:
        if key not in parameters:
            continue
        parameter_sheet.append([
            PARAMETER_EXPORT_NAMES.get(key, key),
            _excel_safe(parameters.get(key)),
            key,
        ])

    for key in ("_parameter_source", "_parameter_recognized", "_parameter_ignored", "_web_overrides"):
        import_sheet.append([key, _excel_safe(parameters.get(key, ""))])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for current_sheet in workbook.worksheets:
        for cell in current_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for cells in current_sheet.columns:
            current_sheet.column_dimensions[get_column_letter(cells[0].column)].width = min(
                max(max(len(str(cell.value or "")) for cell in cells) + 2, 10),
                45,
            )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_data_uri(results: list[dict[str, Any]], meta: dict[str, Any]) -> str:
    encoded = base64.b64encode(build_export(results, meta)).decode("ascii")
    return "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64," + encoded
