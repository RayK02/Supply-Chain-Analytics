from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook

from webapp import app
from app.inventory_analysis import calculate_analysis
from app.inventory_utils import normalize_article_key, round_up_to_vpe
from inventory import analyse_workbook, build_export


def test_article_normalization_preserves_erp_identifiers():
    assert normalize_article_key(602) == "000602"
    assert normalize_article_key(100602.0) == "100602"
    assert normalize_article_key("ABC") == "ABC"
    assert normalize_article_key("ART-991202") == "ART-991202"
    assert normalize_article_key("10.9912") == "10.9912"
    assert normalize_article_key("991204-A") == "991204-A"
    assert normalize_article_key("X99") == "X99"
    assert normalize_article_key("1e6") == "1E6"
    assert normalize_article_key(991207.5) is None


def test_vpe_rounding():
    assert round_up_to_vpe(101, 20) == 120
    assert round_up_to_vpe(100, 20) == 100


def test_abcxyz_calculation():
    rows = [
        {"article_key": "100001", "booking_date": date(2026, 1, 31), "document_type": "Verkaufslieferung", "quantity": -80},
        {"article_key": "100001", "booking_date": date(2026, 2, 28), "document_type": "Verkaufslieferung", "quantity": -80},
        {"article_key": "100001", "booking_date": date(2026, 3, 31), "document_type": "Verkaufslieferung", "quantity": -80},
        {"article_key": "100002", "booking_date": date(2026, 3, 31), "document_type": "Verkaufslieferung", "quantity": -20},
    ]
    results, meta = calculate_analysis(rows, vpe_by_article={"100001": 20})
    by_key = {row["article_key"]: row for row in results}
    assert by_key["100001"]["abc_class"] == "A"
    assert by_key["100001"]["xyz_class"] == "X"
    assert by_key["100001"]["order_quantity"] % 20 == 0
    assert meta["overall_sales"] == 260


def sample_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Artikelposten"
    ws.append(["Artikelnummer", "Beschreibung", "Buchungsdatum", "Belegart", "Menge"])
    for month, day, qty in ((1, 31, -10), (2, 28, -12), (3, 31, -11)):
        ws.append([100001, "Testartikel", date(2026, month, day), "Verkaufslieferung", qty])
    master = wb.create_sheet("Artikel_Stamm")
    master.append(["Artikelnummer", "Beschreibung"])
    master.append([100001, "Testartikel"])
    vpe = wb.create_sheet("VPE")
    vpe.append(["Artikelnummer", "VPE"])
    vpe.append([100001, 10])
    current = wb.create_sheet("Lagerhaltungsdaten_IST")
    current.append(["Artikelnummer", "Lagerortcode", "Mindestbestand", "Bestellmenge"])
    current.append([100001, "FERTIG", 20, 10])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def test_workbook_analysis_and_export():
    results, meta = analyse_workbook(sample_workbook())
    assert len(results) == 1
    assert results[0]["article_key"] == "100001"
    assert results[0]["vpe"] == 10
    exported = build_export(results, meta)
    workbook = load_workbook(BytesIO(exported), data_only=True)
    assert workbook.sheetnames == [
        "ABC_Analyse",
        "Zusammenfassung",
        "IST_Quellen",
        "Importbericht",
        "Parameter",
    ]


def test_export_neutralizes_formula_injection():
    results, meta = analyse_workbook(sample_workbook())
    results[0]["description"] = '=HYPERLINK("https://example.invalid","Öffnen")'
    exported = build_export(results, meta)
    workbook = load_workbook(BytesIO(exported), data_only=False)
    assert workbook["ABC_Analyse"]["B2"].value.startswith("'=")
    assert workbook["ABC_Analyse"]["B2"].data_type == "s"


def test_health_and_invalid_upload():
    client = app.test_client()
    assert client.get("/health").status_code == 200
    assert client.get("/analyze").status_code == 303
    assert client.post("/analyze", data={}).status_code == 400
    assert client.post("/analyze", data={"file": (BytesIO(b"x"), "test.csv")}, content_type="multipart/form-data").status_code == 400
