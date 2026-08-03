from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook, load_workbook

from analysis_service import analyse_workbook
from app.inventory_analysis import calculate_analysis
from inventory import build_export, params
from webapp import app


def workbook_stream(workbook: Workbook) -> io.BytesIO:
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def test_audit_alphanumeric_article_numbers_survive_full_workbook_flow():
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge", "Beschreibung"])
    article_numbers = ["ART-991202", "10.9912", "991204-A", "X99", "abc-100"]
    for article in article_numbers:
        sales.append([article, date(2026, 6, 30), "Verkaufslieferung", -10, article])

    results, meta = analyse_workbook(
        workbook_stream(workbook),
        analysis_settings={"months_average": 1, "xyz_months": 3},
    )

    assert {row["article_display"] for row in results} == set(article_numbers)
    assert len(results) == len(article_numbers)
    assert meta["sales_import"]["invalid_article_rows"] == 0


def test_audit_power_of_two_decoder_uses_exactly_three_complete_months():
    rows = [
        {
            "article_key": "A-1",
            "article_display": "A-1",
            "booking_date": date(2026, 3, 31),
            "document_type": "Verkaufslieferung",
            "quantity": -128,
        },
        {
            "article_key": "A-1",
            "article_display": "A-1",
            "booking_date": date(2026, 4, 30),
            "document_type": "Verkaufslieferung",
            "quantity": -256,
        },
        {
            "article_key": "A-1",
            "article_display": "A-1",
            "booking_date": date(2026, 5, 31),
            "document_type": "Verkaufslieferung",
            "quantity": -512,
        },
        {
            "article_key": "A-1",
            "article_display": "A-1",
            "booking_date": date(2026, 6, 30),
            "document_type": "Verkaufslieferung",
            "quantity": -1024,
        },
    ]

    results, meta = calculate_analysis(
        rows,
        parameters={
            "analysis_end_date": date(2026, 6, 30),
            "months_average": 3,
            "xyz_months": 3,
        },
    )
    result = results[0]

    assert result["monthly_sales"] == [256, 512, 1024]
    assert result["sales_recent"] == 1792
    assert result["average_month"] == 1792 / 3
    assert result["minimum_stock"] == 1792
    assert meta["months_average_used"] == 3


def test_audit_own_export_parameter_sheet_roundtrips():
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge"])
    sales.append(["A-1", date(2026, 3, 31), "Verkaufslieferung", -100])

    parameter_sheet = workbook.create_sheet("Parameter")
    parameter_sheet.append(["Parameter", "Wert"])
    parameter_sheet.append(["Durchschnittsmonate", 3])
    parameter_sheet.append(["XYZ-Monate", 12])
    parameter_sheet.append(["Mindestfaktor A", 10])
    parameter_sheet.append(["Startdatum", date(2026, 1, 1)])
    parameter_sheet.append(["Enddatum", date(2026, 3, 31)])

    results, meta = analyse_workbook(workbook_stream(workbook))
    assert results[0]["minimum_stock"] == 1000

    exported = build_export(results, meta)
    exported_workbook = load_workbook(io.BytesIO(exported), data_only=True)
    roundtrip = params(exported_workbook)

    assert roundtrip["months_average"] == 3
    assert roundtrip["xyz_months"] == 12
    assert roundtrip["minimum_factor_a"] == 10
    assert roundtrip["analysis_start_date"] == date(2026, 1, 1)
    assert roundtrip["analysis_end_date"] == date(2026, 3, 31)


def test_audit_negative_credit_quantity_is_a_return_correction():
    rows = [
        {
            "article_key": "A-1",
            "booking_date": date(2026, 6, 30),
            "document_type": "Verkaufslieferung",
            "quantity": -1200,
        },
        {
            "article_key": "A-1",
            "booking_date": date(2026, 6, 30),
            "document_type": "Verkaufsgutschrift",
            "quantity": -360,
        },
    ]

    results, meta = calculate_analysis(
        rows,
        parameters={"months_average": 1, "xyz_months": 3},
    )
    result = results[0]

    assert result["sales_gross"] == 1200
    assert result["returns_total"] == -360
    assert result["sales_total"] == 1560
    assert meta["overall_returns"] == -360
    assert meta["overall_sales"] == 1560


def test_result_page_has_no_base64_export_and_states_public_access():
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge"])
    sales.append(["A-1", date(2026, 6, 30), "Verkaufslieferung", -100])

    client = app.test_client()
    response = client.post(
        "/analyze",
        data={
            "analysis_file": (workbook_stream(workbook), "analysis.xlsx"),
            "months_average": "1",
            "xyz_months": "3",
            "analysis_start_date": "",
            "analysis_end_date": "",
            "output_mode": "view",
        },
        content_type="multipart/form-data",
    )
    text = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64" not in text
    assert "Öffentliche Webversion ohne Zugriffsschutz" in text
    assert "Für XLSX: Dateien oben erneut auswählen" in text
    assert "Der Browser gibt die Dateiauswahl nach dem Seitenwechsel nicht zurück" in text
