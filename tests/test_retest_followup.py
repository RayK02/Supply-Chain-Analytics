from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from analysis_service import analyse_workbook
from app.inventory_analysis import calculate_analysis
from inventory import build_export, params
from webapp import app


def workbook_stream(workbook: Workbook) -> io.BytesIO:
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def simple_analysis_workbook(article: str = "A-1") -> io.BytesIO:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge", "Beschreibung"])
    sales.append([article, date(2026, 6, 30), "Verkaufslieferung", -100, "Testartikel"])
    return workbook_stream(workbook)


def test_parameter_export_roundtrip_accepts_german_and_technical_keys():
    workbook = Workbook()
    parameter_sheet = workbook.active
    parameter_sheet.title = "Parameter"
    parameter_sheet.append(["Parameter", "Wert"])
    parameter_sheet.append(["months_average", 4])
    parameter_sheet.append(["XYZ-Monate", 9])
    parameter_sheet.append(["Startdatum", date(2026, 1, 1)])
    parameter_sheet.append(["analysis_end_date", date(2026, 6, 30)])
    parameter_sheet.append(["Rückgabebelegarten", "Verkaufsgutschrift"])
    parsed = params(workbook)
    assert parsed["months_average"] == 4
    assert parsed["xyz_months"] == 9
    assert parsed["analysis_start_date"] == date(2026, 1, 1)
    assert parsed["analysis_end_date"] == date(2026, 6, 30)
    assert parsed["return_document_types"] == "Verkaufsgutschrift"

    exported = build_export([], {"parameters": parsed})
    roundtrip = params(load_workbook(io.BytesIO(exported), data_only=True))
    assert roundtrip["months_average"] == 4
    assert roundtrip["xyz_months"] == 9
    assert roundtrip["analysis_start_date"] == date(2026, 1, 1)
    assert roundtrip["analysis_end_date"] == date(2026, 6, 30)
    assert roundtrip["return_document_types"] == "Verkaufsgutschrift"


def test_parameter_max_import_rows_is_enforced():
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge"])
    for index in range(3):
        sales.append([f"A-{index}", date(2026, 6, 30), "Verkaufslieferung", -10])
    parameter_sheet = workbook.create_sheet("Parameter")
    parameter_sheet.append(["Parameter", "Wert"])
    parameter_sheet.append(["Maximale Importzeilen", 2])

    with pytest.raises(ValueError, match="überschreitet 2 Datenzeilen"):
        analyse_workbook(workbook_stream(workbook))


def test_average_window_before_data_start_uses_only_covered_months():
    rows = []
    for month in range(8, 13):
        rows.append({
            "article_key": "A-1",
            "article_display": "A-1",
            "booking_date": date(2025, month, 15),
            "document_type": "Verkaufslieferung",
            "quantity": -100,
        })
    for month in range(1, 7):
        rows.append({
            "article_key": "A-1",
            "article_display": "A-1",
            "booking_date": date(2026, month, 15),
            "document_type": "Verkaufslieferung",
            "quantity": -100,
        })

    results, meta = calculate_analysis(rows, parameters={
        "analysis_end_date": date(2026, 6, 30),
        "months_average": 12,
        "xyz_months": 12,
    })
    assert meta["months_average_used"] == 11
    assert results[0]["average_month"] == 100
    assert any("11 statt 12" in warning for warning in meta["warnings"])


def test_missing_vpe_is_grouped_and_does_not_force_manual_status():
    results, meta = analyse_workbook(simple_analysis_workbook(), analysis_settings={
        "months_average": 1,
        "xyz_months": 3,
    })
    assert meta["missing_vpe_count"] == 1
    assert meta["validation_count"] == 0
    assert results[0]["validation_messages"] == []
    assert results[0]["overall_status"] == "missing"
    assert any("Für 1 Artikel fehlt die VPE" in warning for warning in meta["warnings"])


def test_negative_return_quantity_reverses_a_previous_return():
    rows = [
        {"article_key": "A-1", "booking_date": date(2026, 6, 30), "document_type": "Verkaufslieferung", "quantity": -100},
        {"article_key": "A-1", "booking_date": date(2026, 6, 30), "document_type": "Verkaufsgutschrift", "quantity": 30},
        {"article_key": "A-1", "booking_date": date(2026, 6, 30), "document_type": "Verkaufsgutschrift", "quantity": -10},
    ]
    results, _ = calculate_analysis(rows, parameters={"months_average": 1, "xyz_months": 3})
    assert results[0]["sales_gross"] == 100
    assert results[0]["returns_total"] == 20
    assert results[0]["sales_total"] == 80


def test_article_display_preserves_original_case_in_ui_and_export():
    results, meta = analyse_workbook(simple_analysis_workbook("1e6"), analysis_settings={
        "months_average": 1,
        "xyz_months": 3,
    })
    assert results[0]["article_key"] == "1E6"
    assert results[0]["article_display"] == "1e6"
    exported = build_export(results, meta)
    workbook = load_workbook(io.BytesIO(exported), data_only=True)
    assert workbook["ABC_Analyse"]["A2"].value == "1e6"


def test_export_neutralizes_formula_injection_after_leading_whitespace():
    results, meta = analyse_workbook(simple_analysis_workbook(), analysis_settings={
        "months_average": 1,
        "xyz_months": 3,
    })
    results[0]["description"] = "  =1+1"
    exported = build_export(results, meta)
    workbook = load_workbook(io.BytesIO(exported), data_only=False)
    cell = workbook["ABC_Analyse"]["B2"]
    assert cell.value.startswith("'  =")
    assert cell.data_type == "s"


def test_csp_uses_external_script_and_direct_download_returns_xlsx():
    client = app.test_client()
    index_response = client.get("/")
    csp = index_response.headers["Content-Security-Policy"]
    assert "'unsafe-inline'" not in csp
    assert "/static/app.js" in index_response.get_data(as_text=True)

    response = client.post(
        "/analyze",
        data={
            "analysis_file": (simple_analysis_workbook(), "analysis.xlsx"),
            "months_average": "1",
            "xyz_months": "3",
            "analysis_start_date": "",
            "analysis_end_date": "",
            "output_mode": "download",
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert response.data.startswith(b"PK")
    assert "attachment" in response.headers["Content-Disposition"]
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_ignored_parameter_is_visible_on_result_page():
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge"])
    sales.append(["A-1", date(2026, 6, 30), "Verkaufslieferung", -10])
    parameter_sheet = workbook.create_sheet("Parameter")
    parameter_sheet.append(["Parameter", "Wert"])
    parameter_sheet.append(["Unbekannter Schlüssel", 123])

    client = app.test_client()
    response = client.post(
        "/analyze",
        data={
            "analysis_file": (workbook_stream(workbook), "analysis.xlsx"),
            "months_average": "1",
            "xyz_months": "3",
            "analysis_start_date": "",
            "analysis_end_date": "",
        },
        content_type="multipart/form-data",
    )
    text = response.get_data(as_text=True)
    assert response.status_code == 200
    assert "1 Parameter wurde nicht erkannt" in text
    assert "Unbekannter Schlüssel" in text
