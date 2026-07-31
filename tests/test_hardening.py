from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import Workbook

from analysis_service import analyse_workbook
from app.inventory_analysis import calculate_analysis
from inventory import current_values, params, validate_xlsx_archive


def workbook_stream(workbook: Workbook) -> io.BytesIO:
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def base_workbook(rows: list[list], *, master_articles: list[str] | None = None) -> io.BytesIO:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge", "Beschreibung"])
    for row in rows:
        sales.append(row)
    if master_articles:
        master = workbook.create_sheet("Artikel_Stamm")
        master.append(["Artikelnr.", "Beschreibung"])
        for article in master_articles:
            master.append([article, f"Artikel {article}"])
    return workbook_stream(workbook)


def test_artikelposten_is_never_used_as_embedded_ist_source():
    stream = base_workbook([
        ["ART-1", date(2026, 6, 30), "Verkaufslieferung", -100, "Artikel 1"],
    ])
    results, meta = analyse_workbook(stream, analysis_settings={"months_average": 1})
    assert len(results) == 1
    assert results[0]["order_quantity_current"] is None
    assert results[0]["current_source_file"] == ""
    assert meta["current_articles"] == 0
    assert meta["current_sources"] == []


def test_alphanumeric_and_separator_article_numbers_are_retained():
    article_numbers = ["ART-991202", "10.9912", "991204-A", "X99", "1e6"]
    rows = [
        [article, date(2026, 6, 30), "Verkaufslieferung", -10, article]
        for article in article_numbers
    ]
    results, meta = analyse_workbook(base_workbook(rows), analysis_settings={"months_average": 1})
    assert {row["article_key"] for row in results} == {
        "ART-991202", "10.9912", "991204-A", "X99", "1E6"
    }
    assert meta["sales_import"]["invalid_article_rows"] == 0


def test_master_data_without_movements_does_not_create_phantom_rows():
    stream = base_workbook(
        [["SOLD", date(2026, 6, 30), "Verkaufslieferung", -10, "Verkauft"]],
        master_articles=["SOLD", "INACTIVE-1", "INACTIVE-2"],
    )
    results, _ = analyse_workbook(stream, analysis_settings={"months_average": 1})
    assert [row["article_key"] for row in results] == ["SOLD"]


def test_returns_and_positive_reversals_reduce_net_demand():
    rows = [
        {"article_key": "A-1", "booking_date": date(2026, 6, 30), "document_type": "Verkaufslieferung", "quantity": -100},
        {"article_key": "A-1", "booking_date": date(2026, 6, 30), "document_type": "Verkaufslieferung", "quantity": 20},
        {"article_key": "A-1", "booking_date": date(2026, 6, 30), "document_type": "Verkaufsrücksendung", "quantity": 10},
    ]
    results, meta = calculate_analysis(rows, parameters={"months_average": 1})
    assert results[0]["sales_gross"] == 100
    assert results[0]["returns_total"] == 30
    assert results[0]["sales_total"] == 70
    assert meta["overall_sales"] == 70


def test_parameter_sheet_is_applied_when_web_override_is_blank():
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge"])
    sales.append(["A-1", date(2026, 5, 31), "Verkaufslieferung", -10])
    sales.append(["A-1", date(2026, 6, 30), "Verkaufslieferung", -30])
    parameter_sheet = workbook.create_sheet("Parameter")
    parameter_sheet.append(["Parameter", "Wert"])
    parameter_sheet.append(["Durchschnittsmonate", 2])
    parameter_sheet.append(["Mindestfaktor A", 4])

    stream = workbook_stream(workbook)
    results, meta = analyse_workbook(stream, analysis_settings={})
    assert meta["months_average"] == 2
    assert results[0]["average_month"] == 20
    assert results[0]["minimum_stock"] == 80
    assert "Durchschnittsmonate" in meta["parameters"]["_parameter_recognized"]


def test_plausibility_checks_flag_minimum_above_maximum():
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge"])
    sales.append(["A-1", date(2026, 6, 30), "Verkaufslieferung", -100])
    current = workbook.create_sheet("Lagerhaltungsdaten_IST")
    current.append(["Artikelnr.", "Lagerortcode", "Minimalbestand", "Bestellmenge", "Maximalbestand"])
    current.append(["A-1", "FERTIG", 20, 10, 10])

    results, meta = analyse_workbook(workbook_stream(workbook), analysis_settings={"months_average": 1})
    assert results[0]["overall_status"] == "manual"
    assert "Minimalbestand IST liegt über dem Maximalbestand IST" in results[0]["validation_messages"]
    assert meta["validation_count"] == 1


def test_generic_menge_column_is_not_bestellmenge():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Unbekannt"
    sheet.append(["Artikelnr.", "Menge"])
    sheet.append(["A-1", -100])
    with pytest.raises(ValueError, match="Keine erkennbare Lagerhaltungsdatenliste"):
        current_values(
            workbook,
            params(workbook),
            source_file="unknown.xlsx",
            required=True,
            scan_all_sheets=True,
        )


def test_invalid_xlsx_archive_is_rejected_cleanly():
    with pytest.raises(ValueError, match="gültige XLSX"):
        validate_xlsx_archive(io.BytesIO(b"not-an-xlsx"), "broken.xlsx")
