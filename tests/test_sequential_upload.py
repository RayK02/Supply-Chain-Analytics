from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook, load_workbook

import analysis_service
from analysis_service import (
    add_current_workbook,
    finalize_prepared_analysis,
    prepare_analysis_workbook,
)
from analysis_token import decode_analysis_token, encode_analysis_token
from webapp import app


def workbook_stream(workbook: Workbook) -> io.BytesIO:
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def analysis_workbook() -> Workbook:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge", "Beschreibung"])
    sales.append(["A-100", date(2026, 6, 30), "Verkaufslieferung", -100, "Testartikel"])
    return workbook


def current_workbook() -> Workbook:
    workbook = Workbook()
    current = workbook.active
    current.title = "Lagerhaltungsdatenübersicht"
    current.append([
        "Artikelnr.",
        "Lagerortcode",
        "Lagerbestand",
        "Minimalbestand",
        "Bestellmenge",
        "Maximalbestand",
    ])
    current.append(["A-100", "KELLERA", 25, 300, 50, 500])
    return workbook


def prepare_token(client) -> str:
    response = client.post(
        "/api/prepare-analysis",
        data={
            "analysis_file": (workbook_stream(analysis_workbook()), "analysis.xlsx"),
            "months_average": "1",
            "xyz_months": "3",
            "analysis_start_date": "",
            "analysis_end_date": "",
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    return response.get_json()["token"]


def test_analysis_token_roundtrip_preserves_uploaded_data_without_precomputed_results():
    payload = prepare_analysis_workbook(
        workbook_stream(analysis_workbook()),
        {"months_average": 1, "xyz_months": 3},
        source_filename="analysis.xlsx",
    )
    payload["display_settings"] = {
        "months_average": 1,
        "xyz_months": 3,
        "analysis_start_date": "",
        "analysis_end_date": "",
    }

    decoded = decode_analysis_token(encode_analysis_token(payload))

    assert decoded["source_filename"] == "analysis.xlsx"
    assert decoded["sales"][0][1] == "A-100"
    assert decoded["sales"][0][3] == date(2026, 6, 30)
    assert "results" not in decoded
    assert "meta" not in decoded


def test_calculation_runs_once_only_after_all_uploads(monkeypatch):
    calls: list[int] = []
    real_calculate = analysis_service.calculate_analysis

    def tracked_calculate(*args, **kwargs):
        calls.append(1)
        return real_calculate(*args, **kwargs)

    monkeypatch.setattr(analysis_service, "calculate_analysis", tracked_calculate)

    payload = prepare_analysis_workbook(
        workbook_stream(analysis_workbook()),
        {"months_average": 1, "xyz_months": 3},
        source_filename="analysis.xlsx",
    )
    assert calls == []

    add_current_workbook(payload, "current.xlsx", workbook_stream(current_workbook()))
    assert calls == []

    results, meta = finalize_prepared_analysis(payload)
    assert calls == [1]
    assert results[0]["stock_current"] == 25
    assert meta["calculation_stage"] == "after_all_uploads"


def test_sequential_api_adds_current_workbook_and_renders_result():
    client = app.test_client()
    token = prepare_token(client)

    current_response = client.post(
        "/api/add-current",
        data={
            "analysis_token": token,
            "current_file": (workbook_stream(current_workbook()), "current.xlsx"),
        },
        content_type="multipart/form-data",
    )
    assert current_response.status_code == 200

    final_response = client.post(
        "/api/finalize-analysis",
        data={
            "analysis_token": current_response.get_json()["token"],
            "output_mode": "view",
        },
        content_type="multipart/form-data",
    )
    text = final_response.get_data(as_text=True)

    assert final_response.status_code == 200
    assert "current.xlsx" in text
    assert "A-100" in text
    assert "KELLERA" in text
    assert "Die Dateien werden nacheinander verarbeitet" in text


def test_sequential_api_download_uses_existing_python_export():
    client = app.test_client()
    token = prepare_token(client)
    current_response = client.post(
        "/api/add-current",
        data={
            "analysis_token": token,
            "current_file": (workbook_stream(current_workbook()), "current.xlsx"),
        },
        content_type="multipart/form-data",
    )

    download = client.post(
        "/api/finalize-analysis",
        data={
            "analysis_token": current_response.get_json()["token"],
            "output_mode": "download",
        },
        content_type="multipart/form-data",
    )

    assert download.status_code == 200
    assert download.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    exported = load_workbook(io.BytesIO(download.data), data_only=True)
    assert exported.sheetnames == ["ABC_Analyse", "Zusammenfassung", "IST_Quellen", "Importbericht", "Parameter"]
    assert exported["ABC_Analyse"]["A2"].value == "A-100"
    assert exported["ABC_Analyse"]["K2"].value == 25


def test_invalid_analysis_token_is_rejected_cleanly():
    client = app.test_client()
    response = client.post(
        "/api/finalize-analysis",
        data={"analysis_token": "not-a-token", "output_mode": "view"},
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "temporäre Analysestatus" in response.get_json()["error"]


def test_index_explains_single_file_limit_and_has_progress_region():
    client = app.test_client()
    response = client.get("/")
    text = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "jede einzelne Datei darf maximal ca. 4,1 MB gross sein" in text
    assert 'id="clientStatus"' in text
    assert "zusammen zu gross" not in text
