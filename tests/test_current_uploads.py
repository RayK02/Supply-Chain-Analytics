from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook, load_workbook

from app.inventory_utils import DEFAULT_PARAMETERS
from inventory import analyse_workbook, current_values
from webapp import app


def workbook_bytes(workbook: Workbook) -> io.BytesIO:
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def make_analysis_workbook() -> io.BytesIO:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Artikelposten"
    sales.append(["Artikelnr.", "Buchungsdatum", "Belegart", "Menge", "Beschreibung"])
    for day, quantity in ((date(2026, 5, 31), -30), (date(2026, 6, 30), -40), (date(2026, 7, 31), -50)):
        sales.append(["000040", day, "Verkaufslieferung", quantity, "microSD 4GB ohne Adapter"])

    articles = workbook.create_sheet("Artikel_Stamm")
    articles.append(["Artikelnr.", "Beschreibung"])
    articles.append(["000040", "microSD 4GB ohne Adapter"])

    vpe = workbook.create_sheet("VPE")
    vpe.append(["Artikelnr.", "VPE"])
    vpe.append(["000040", 100])

    embedded = workbook.create_sheet("Lagerhaltungsdaten_IST")
    embedded.append(["Artikelnr.", "Lagerortcode", "Mindestbestand", "Bestellmenge", "Maximalbestand"])
    embedded.append(["000040", "FERTIG", 1, 1, 1])
    return workbook_bytes(workbook)


def make_current_workbook() -> io.BytesIO:
    workbook = Workbook()
    current = workbook.active
    current.title = "Lagerhaltungsdatenübersicht"
    current.append([
        "Artikelnr.", "Variantencode", "Lagerortcode", "Beschaffungsmethode",
        "Beschreibung", "Lagerbestand", "Minimalbestand", "Bestellmenge", "Maximalbestand",
    ])
    current.append(["000040", "", "FERTIG", "Einkauf", "microSD 4GB ohne Adapter", 520, 90, 500, 1])
    return workbook_bytes(workbook)


def test_current_export_aliases_are_recognized():
    values, info = current_values(
        load_workbook(make_current_workbook(), data_only=True, read_only=True),
        dict(DEFAULT_PARAMETERS),
        source_file="current.xlsx",
        required=True,
    )
    assert info["sheets"] == ["Lagerhaltungsdatenübersicht"]
    assert values["000040"]["stock_current"] == 520
    assert values["000040"]["minimum_stock_current"] == 90
    assert values["000040"]["order_quantity_current"] == 500
    assert values["000040"]["maximum_stock_current"] == 1


def test_external_current_file_overrides_embedded_values():
    results, meta = analyse_workbook(
        make_analysis_workbook(),
        [("current.xlsx", make_current_workbook())],
    )
    result = next(row for row in results if row["article_key"] == "000040")
    assert result["stock_current"] == 520
    assert result["minimum_stock_current"] == 90
    assert result["order_quantity_current"] == 500
    assert result["current_source_file"] == "current.xlsx"
    assert meta["external_current_files"] == 1


def test_web_route_accepts_multiple_current_files():
    client = app.test_client()
    response = client.post(
        "/analyze",
        data={
            "analysis_file": (make_analysis_workbook(), "analysis.xlsx"),
            "current_files": [
                (make_current_workbook(), "current-1.xlsx"),
                (make_current_workbook(), "current-2.xlsx"),
            ],
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"current-1.xlsx" in response.data
    assert b"current-2.xlsx" in response.data
    assert b"Lagerbestand IST" in response.data


def test_invalid_current_workbook_returns_clear_error():
    invalid = Workbook()
    invalid.active.append(["Unbekannt", "Wert"])
    client = app.test_client()
    response = client.post(
        "/analyze",
        data={
            "analysis_file": (make_analysis_workbook(), "analysis.xlsx"),
            "current_files": [(workbook_bytes(invalid), "invalid.xlsx")],
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 400
    assert "Keine erkennbare Lagerhaltungsdatenliste" in response.get_data(as_text=True)
